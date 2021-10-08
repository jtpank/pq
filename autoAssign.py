import openpyxl as opx
import os
import subprocess
import csv
import time
import re
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from datetime import date, timedelta
#USAGE
#   1. Headless login to the url
#   2. Download csv from specified date
#   3. Convert CSV to XLSX | NOT NECESSARY --> then 4b becomes: Parse CSV to a dict
#   4a. Determine if the test is "complete" vs "new" (only want the complete tests that have no result)
#   4b. Parse XLSX to a dict and store pairs "name": "test date"
#   5a. Use api to verify on google drive that the dict pairs HAVE A NEGATIVE TEST
#   5b. Loop through each pair in dict and assign result using selenium




#Selenium chromedriver
options = Options()
options.headless = True
options.add_experimental_option("prefs", {
  "download.default_directory": "/root/pythonProject/downloads",
  'download.prompt_for_download': False,
  'download.directory_upgrade': True,
  'safebrowsing.enabled': False,
  'safebrowsing.disable_download_protection': True,
})
options.add_argument('--disable-extensions')
options.add_argument('--headless')
options.add_argument('--disable-gpu')
options.add_argument('--no-sandbox')
CHROMEDRIVER_PATH = './chromedriver'

_FILEPATH = "../downloads/exportRecentXLSX.xlsx"
_PATHCSV = "../downloads/exportRecent.csv"

#   1. Headless login to the url
#   2. Download csv from specified date, headless
# TODO: specify date
#DigitalOcean PW:
#Email: 
#Project: PQ
#Authentication password: 
URL = ""
USERNAME = ""
PASSWORD = ""
#search 2 days from today
today = date.today().strftime('%m/%d/%y')
yesterday = date.today() - timedelta(days=2)
inputTime = yesterday.strftime('%m/%d/%y')
def pull_current_CSV_PQ(url):
    os.system("rm -rf ../downloads/")
    os.system("mkdir ../downloads/")
    #WEBROWSER BEGIN
    driver = webdriver.Chrome(CHROMEDRIVER_PATH, options=options)
    driver.get(URL)
    #wait for page to load before searching by xpath
    driver.implicitly_wait(8)
    #chrome inspector: $x("//login-form/form/div/div[@class='email-box']/waf-email-input/form-control-container/div/div/div")
    #LOGIN BEGIN
    user = driver.find_element_by_xpath("//login-form/form/div/div[@class='email-box']/waf-email-input/form-control-container/div/div/div/input[@type='email']")
    user.send_keys(USERNAME)
    pw = driver.find_element_by_xpath("//login-form/form/div/div[@class='password-box']/waf-password-input/form-control-container/div/div/input[@type='password']")
    pw.send_keys(PASSWORD)
    driver.find_element_by_xpath("//login-form/form/div/div[@class='login-actions']").click()
    driver.implicitly_wait(8)
    #LOGIN COMPLETE

    #ADVANCED SEARCH BEGIN
    #click on drop down span
    driver.find_element_by_xpath("//body/app-root/div/div/div[@class='page-content-wrapper']/div/div/order-search/tcm-main-search-layout/tcm-main-frame/div/div/div[2]/div/search-common-commands/div/button[@class='btn btn-link advanced-search-link']").click()
    driver.implicitly_wait(5)
    advSearch = driver.find_element_by_xpath("//body/app-root/div/div/div[@class='page-content-wrapper']/div/div/order-search/tcm-main-search-layout/tcm-main-frame/div/div[@class='body-container scrollable-content']/div/div/div/div[2]/advanced-search/div/form/div/div/waf-date-input/form-control-container/div/div/div/div/input")
    advSearch.send_keys(inputTime)
    driver.implicitly_wait(5)
    driver.find_element_by_xpath("//body/app-root/div/div/div[@class='page-content-wrapper']/div/div/order-search/tcm-main-search-layout/tcm-main-frame/div/div[@class='body-container scrollable-content']/div/div/div/div[2]/advanced-search/div/div[@class='actions']/button[@class='btn btn-complete']").click()
    driver.implicitly_wait(5)
    #ADVANCED SEARCH END

    #DOWNLOAD CSV BEGIN
    driver.command_executor._commands["send_command"] = ("POST", '/session/$sessionId/chromium/send_command')
    params = {'cmd': 'Page.setDownloadBehavior', 'params': {'behavior': 'allow', 'downloadPath': "C:\\Users\\jtpan\\OneDrive\\Desktop\\bickmanDisneyCovid\\downloads"}}
    command_result = driver.execute("send_command", params)
    print("response from browser:")
    for key in command_result:
        print("result:" + key + ":" + str(command_result[key]))

    export_CSV = driver.find_element_by_class_name("export-to-csv").click()
    time.sleep(4)
    #DOWNLOAD CSV END
    driver.quit()
    #remove spaces from filename
    #rename file
    #in shell script
    subprocess.call(['sh', './removeRename.sh'])

#   3. Convert CSV to XLSX
def convertCSV_XLSX(pathCSV):
    wb = opx.Workbook()
    ws = wb.active
    with open(pathCSV) as f:
        reader = csv.reader(f, delimiter=',')
        for row in reader:
            ws.append([ILLEGAL_CHARACTERS_RE.sub('', _i) for _i in row])
    wb.save("../downloads/exportRecentXLSX.xlsx")

#   4a. Determine if the test is "complete" vs "new" (only want the complete tests that have no result)
#   4b. Parse XLSX to a dict and store pairs "name": "test date"
#TODO: assign to dict
# start from row 2 because of header row in .csv
startRow = 2
#letters are excel columns
firstNameCol = 21 #V
lastNameCol = 22 #W
resultCol = 15 #P
barcodeCol = 7 #H
statusCol = 3
dateCol = 8 #J
FullNameDateDict = {}
NameDateDict_withFullNameTuple = {}
def pullFullNamesCompleteTests():
    wb = opx.load_workbook(_FILEPATH)
    sheet1 = wb.sheetnames[0]
    ws = wb[sheet1]
    for row in range(startRow, ws.max_row+1):
            rowBarCode = ws[row][barcodeCol].value
            rowResult = ws[row][resultCol].value
            rowStatus = ws[row][statusCol].value
            if not rowResult and rowBarCode and rowStatus == 'Complete':    
                rowNames = [cell.value for cell in ws[row][firstNameCol:lastNameCol+1]]
                rowDate = ws[row][dateCol].value
                rowDateStr = str(rowDate).split()[0]
                #store full names for google docs search
                if not rowDateStr[:5] == today[:5]:
                    FullNameDateDict[rowNames[0]+ ' ' + rowNames[1]] = rowDateStr
                #if rowDateStr does not equal to current date then store in dict
                # rowDateStr is stored as 'mm/dd/yy' inputTime is stored as 'mm/dd/yyyy'
                #Handle names of form Firstname = "str1 str2" and remove "str2"
                    # ex: Mark Daniel Quintos --> Mark Quintos
                fullNameDate_tuple = (rowNames[0] + ' ' + rowNames[1], rowDateStr)
                if ' ' in rowNames[0]:
                    firstPart = rowNames[0].split()[0]
                    rowNames[0] = firstPart
                if ' ' in rowNames[1]:
                    firstPartLastName = rowNames[1].split()[0]
                    rowNames[1] = firstPartLastName
                if not rowDateStr[:5] == today[:5]:
                    rowFullName = rowNames[0] + ' ' + rowNames[1]
                    NameDateDict_withFullNameTuple[rowFullName] = fullNameDate_tuple
    with open("names.txt", 'w') as f:
        for name in NameDateDict_withFullNameTuple:
            f.write(name + '\t\t' + NameDateDict_withFullNameTuple[name][0] + '\t\t' + NameDateDict_withFullNameTuple[name][1] + '\n')
        f.close()


#5a.
def verifyNegativeTests(dictIn):
    
    #SETUP for API calls
    creds = None
    if os.path.exists('./googleAPI/token.json'):
        creds = Credentials.from_authorized_user_file('./googleAPI/token.json', SCOPES)
    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                './googleAPI/credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        # Save the credentials for the next run
        with open('./googleAPI/token.json', 'w') as token:
            token.write(creds.to_json())
    service = build(API_NAME, API_VERSION, credentials=creds)
    #FINISH SETUP for API calls

    #QUERY AND DOWNLOAD PDFS
    page_token = None
    fileIDdict = {}
    for person in dictIn:
        fullname = dictIn[person][0]
        dateUnmod = dictIn[person][1]
        # restructure date to be yyyy-mm-dd FROM mm-dd-yyyy
        date = dateUnmod[6:]+ '-' + dateUnmod[:5]
        while True:
            query = "name contains '{0}' and createdTime >= '{1}T00:00:00'".format(fullname, date)
            response = service.files().list(
                #q="name contains 'josh randall'",
                q=query,
                fields='nextPageToken, files(id, name)',
                pageToken=page_token
            ).execute()
            for file in response.get('files', []):
                # Process change
                fileIDdict[fullname] = file.get('id')
                print('Found file and stored in dict: {0} {1}'.format(file.get('name'), file.get('id')))
            page_token = response.get('nextPageToken', None)
            if page_token is None:
                break
    #loop through fileIDdict and download pdfs into folder
    for person in fileIDdict:
        request = service.files().get_media(fileId=fileIDdict[person])
        fh = io.BytesIO()
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while done is False:
            status, done = downloader.next_chunk()
            print("Download {}.".format(int(status.progress() * 100)))
        #file header
        fh.seek(0)
        #save to local drive
        pathStr = "./pdfs/{0}.pdf".format(person)
        with open(os.path.join(pathStr), 'wb') as f:
            f.write(fh.read())
            f.close()
    #FINISH QUERY AND DOWNLOAD PDFS

    #PARSE PDFS and confirm result, name, date, and show
    for person in dictIn:
        pathStr = "./pdfs/{0}.pdf".format(person)
        pdf = PdfFileReader(pathStr)
        pageObj = pdf.getPage(0)
        txt = pageObj.extractText()
        with open('output.txt', 'w') as f:
            f.write(txt)
            f.close()
        lines = []
        with open('output.txt') as f:
            lines = [line.rstrip() for line in f]
            f.close()
        os.system("rm -rf output.txt")
        # Ordering Physician: 
        # Collection Date:
        # Patient Name:
        # Result
        testObj = {'Name': '', 'CollectionDate': '', 'Show': '', 'Result': ''}
        for i in range(0, len(lines)):
            if lines[i] == 'Ordering Physician:':
                strArr = lines[i+1].split('-')
                testObj['Show'] = strArr[1]
            elif lines[i] == 'Collection Date:':
                #format date from m/d/yyyy to mm-dd-yyyy
                strArr = lines[i+1].split('/')
                if len(strArr[0]) == 1:
                    d = '0' + strArr[0]
                    strArr[0] = d
                if len(strArr[1]) == 1:
                    d = '0' + strArr[1]
                    strArr[1] = d
                dateStr = strArr[0] + '-' + strArr[1] + '-' + strArr[2]
                testObj['CollectionDate'] = dateStr
            elif lines[i] == 'Patient Name:':
                strArr = lines[i+1].split(',')
                testObj['Name'] = strArr[1] + ' ' + strArr[0]
            elif lines[i] == 'Result':
                testObj['Result'] = lines[i+1]
            else:
                i+=1
        #Verify 
        if not testObj['Name'] == dictIn[person][0].upper():
            print("Names do not match")
        elif not testObj['CollectionDate'] == dictIn[person][1]:
            print(testObj['CollectionDate'] + '\t' + dictIn[person][1])
            print("Dates do not match")
        elif not testObj['Result'][:8] == 'Negative':
            print("Result is NOT Negative")
        strOut = "{0} tested on {1} for the show {2} and is NEGATIVE \n".format(dictIn[person][0], dictIn[person][1], testObj['Show'])
        print(strOut)
        
        #End Verify
                

#5b. Loop through each pair in dict and assign result using selenium
def assignTestResults(dictIn):
    #WEBROWSER BEGIN
    driver = webdriver.Chrome(CHROMEDRIVER_PATH, options=options)
    driver.get(URL)
    #wait for page to load before searching by xpath
    driver.implicitly_wait(10)
    #chrome inspector: $x("//login-form/form/div/div[@class='email-box']/waf-email-input/form-control-container/div/div/div")
    #LOGIN BEGIN
    user = driver.find_element_by_xpath("//login-form/form/div/div[@class='email-box']/waf-email-input/form-control-container/div/div/div/input[@type='email']")
    user.send_keys(USERNAME)
    pw = driver.find_element_by_xpath("//login-form/form/div/div[@class='password-box']/waf-password-input/form-control-container/div/div/input[@type='password']")
    pw.send_keys(PASSWORD)
    driver.find_element_by_xpath("//login-form/form/div/div[@class='login-actions']").click()
    #LOGIN COMPLETE

    #setup a for loop
    for searchName in dictIn:

        #PATIENT SEARCH BEGIN
        searchInput = driver.find_element_by_xpath("//body/app-root/div/div/div[@class='page-content-wrapper']/div/div/order-search/tcm-main-search-layout/tcm-main-frame/div/div/div[@class='commands']/div/search-common-commands/div/waf-text-input/form-control-container/div/div/div/input[@type='text']")
        searchInput.clear()
        searchInput.send_keys(searchName)
        searchInput.send_keys(Keys.ENTER)
        time.sleep(1)
        
        #wait for search to load results
        #PATIENT SEARCH END

        #RESULT ASSIGNMENT BEGIN
        #verify complete status and date time
        #then click into the patient
        strStatus = ''
        strTestDate = ''
        strPatientName = ''
        i = 1
        while not strStatus == 'Complete' or not strTestDate == dictIn[searchName][1] or not strPatientName == dictIn[searchName][0]:
            s = "//body/app-root/div/div/div[@class='page-content-wrapper']/div/div/order-search/tcm-main-search-layout/tcm-main-frame/div/div[2]/div/div/div/div/search-results/table/tbody/tr[{}]".format(i)
            completeStatus = driver.find_element_by_xpath(s + "/td[1]")
            strStatus = completeStatus.get_attribute('innerText')
            testDate = driver.find_element_by_xpath(s + "/td[2]")
            strTestDate = testDate.get_attribute('innerText').split()[0]
            patientName = driver.find_element_by_xpath(s + "/td[5]")
            strPatientName = patientName.get_attribute('innerText')
            if strStatus == 'Complete' and strTestDate == dictIn[searchName][1] and strPatientName == dictIn[searchName][0]:
                break
            else:
                i += 1
            print(strStatus + '\t' + strTestDate + '\t'+ strPatientName)
            #click into row if verified
        driver.find_element_by_xpath(s).click()    
        #INPUT FORMAT
        #ICD Code: "U07.1 (COVID 19 [confirmed cases])"
        #Result: "Negative"
        #Result Received: "mm/dd/yyyy"
        #Time: "10:00 pm"

        #TODO somehow verify correct show?
        # shows are labeled differently on spreadsheet than portal eg: AHS --> American Horror Story

        #INPUT ICD CODE if empty
        # base path to TEST part of patient sheet
        # //body/app-root/div/div/div[@class='page-content-wrapper']/div/div/order-details/tcm-main-frame/div/div[2]/div/tcm-details-layout/div/div[2]/div/form/collapsible-panel[@name='orderTestDetails']/div/div[2]/div
        # div[1] --> ICD Select
        # div[4] --> Result Select
        # div[5] --> Result Receive Input (date)
        # div[6] --> Time input 

        # for ICD select option IF EMPTY
        # /div[1]/waf-select-input/form-control-container/div/div/div/ng-select/div/div/div[2]"]
        if not driver.find_element_by_xpath("//body/app-root/div/div/div[@class='page-content-wrapper']/div/div/order-details/tcm-main-frame/div/div[2]/div/tcm-details-layout/div/div[2]/div/form/collapsible-panel[@name='orderTestDetails']/div/div[2]/div/div[1]/waf-select-input/form-control-container/div/div/div/ng-select/div/span[@title='Clear all']"):
            icdCodeEntry = driver.find_element_by_xpath("//body/app-root/div/div/div[@class='page-content-wrapper']/div/div/order-details/tcm-main-frame/div/div[2]/div/tcm-details-layout/div/div[2]/div/form/collapsible-panel[@name='orderTestDetails']/div/div[2]/div/div[1]/waf-select-input/form-control-container/div/div/div/ng-select/div/div/div[2]/input")
            icdCodeEntry.send_keys("U07.1 (COVID 19 [confirmed cases])")
            icdCodeEntry.send_keys(Keys.ENTER)
    
        resultEntry = driver.find_element_by_xpath("//body/app-root/div/div/div[@class='page-content-wrapper']/div/div/order-details/tcm-main-frame/div/div[2]/div/tcm-details-layout/div/div[2]/div/form/collapsible-panel[@name='orderTestDetails']/div/div[2]/div/div[4]/waf-select-input/form-control-container/div/div/div/ng-select/div/div/div[2]/input")
        resultEntry.send_keys("Negative")
        resultEntry.send_keys(Keys.ENTER)

        dateEntry = driver.find_element_by_xpath("//body/app-root/div/div/div[@class='page-content-wrapper']/div/div/order-details/tcm-main-frame/div/div[2]/div/tcm-details-layout/div/div[2]/div/form/collapsible-panel[@name='orderTestDetails']/div/div[2]/div/div[5]/waf-date-input/form-control-container/div/div/div/div/input")
        dateEntry.send_keys(dictIn[searchName][1])

        timeEntry = driver.find_element_by_xpath("//body/app-root/div/div/div[@class='page-content-wrapper']/div/div/order-details/tcm-main-frame/div/div[2]/div/tcm-details-layout/div/div[2]/div/form/collapsible-panel[@name='orderTestDetails']/div/div[2]/div/div[6]/waf-time-input/form-control-container/div/div/div/div/input")
        timeEntry.send_keys("10:00 pm")

        selectSaveExit = driver.find_element_by_xpath("//body/app-root/div/div/div[@class='page-content-wrapper']/div/div/order-details/tcm-main-frame/div/div/div[2]/button[@templatetype='SaveExit']")
        print("Result assigned for: " + dictIn[searchName][0] + '\t' + dictIn[searchName][1])
        input("Press enter to continue")
        selectSaveExit.click()
        #RESULT ASSIGNEMENT END
    driver.quit()



def main():
    pull_current_CSV_PQ(URL)
    convertCSV_XLSX(_PATHCSV)
    pullFullNamesCompleteTests()
    
    #verifyNegativeTests(testDict)
    #assignTestResults(NameDateDict_withFullNameTuple)

if __name__ == "__main__":
    main()